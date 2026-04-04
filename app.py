import streamlit as st
import pandas as pd
import io
import calendar
from openpyxl import load_workbook

st.set_page_config(page_title="Roster App", layout="wide")
st.title("Roster Generator")

# MONTH SELECTION
month = st.selectbox("Select Month", list(calendar.month_name)[1:])
year = st.number_input("Select Year", 2024, 2050, 2026)
month_num = list(calendar.month_name).index(month)
days = calendar.monthrange(year, month_num)[1]
st.write("Selected:", month, year, "| Days:", days)

# LOAD FILE
with st.spinner("Loading previous month roster..."):
    df = pd.read_excel("latest_roster.xlsx", skiprows=2)
st.success("File loaded successfully")

# FIND TOTAL COLUMN
total_col_index = None
for i in range(len(df.columns)):
    if "TOTAL" in str(df.columns[i]).upper():
        total_col_index = i
        break

employees = []
prev_duties = {}
emp_rows = {}

for i in range(len(df)):
    name = str(df.iloc[i, 1]).strip()
    if name != "" and name.lower() not in ["nan", "a", "b", "c", "total", "none"]:
        employees.append(name)
        emp_rows[name] = df.iloc[i]

        if total_col_index is not None:
            val = df.iloc[i, total_col_index]
            prev_duties[name] = float(val) if pd.notna(val) else 0
        else:
            prev_duties[name] = 0

# LAST SHIFT
def get_last_shift(row):
    for col in reversed(df.columns):
        val = row[col]
        if pd.notna(val):
            v = str(val).strip().upper()
            if v in ['A', 'B', 'C']:
                return v
    return None

last_shift = {}
for emp in employees:
    last_shift[emp] = get_last_shift(emp_rows[emp])

# DUTY TRACKING
current_duties = {emp: 0 for emp in employees}

# GENERATE ROSTER
roster = {emp: {} for emp in employees}
for d in range(1, days + 1):
    sorted_employees = sorted(
        employees,
        key=lambda x: (current_duties[x] + prev_duties[x])
    )

    workers = sorted_employees[:24]
    off_people = sorted_employees[24:]

    day_roster = {}
    c = b = a = 0

    for emp in workers:
        if c < 8:
            day_roster[emp] = "C"
            c += 1
        elif b < 8:
            day_roster[emp] = "B"
            b += 1
        else:
            if last_shift[emp] == "C":
                day_roster[emp] = "B"
            else:
                day_roster[emp] = "A"
            a += 1

    for emp in off_people:
        day_roster[emp] = "W/O"

    for emp in employees:
        roster[emp][d] = day_roster[emp]

        if day_roster[emp] in ["A", "B", "C"]:
            current_duties[emp] += 1
            last_shift[emp] = day_roster[emp]

# WRITE TO TEMPLATE
wb = load_workbook("Template.xlsx")
ws = wb.active

# FIX HEADER (MERGED CELL SAFE)
for merged_cell in ws.merged_cells.ranges:
    if "1" in str(merged_cell):
        top_left = str(merged_cell).split(":")[0]
        ws[top_left] = f"DUTY ROSTER FOR THE MONTH OF {month.upper()} {year}"
        break

# WRITE DATA
for i, emp in enumerate(employees):
    row = i + 4
    ws.cell(row=row, column=2, value=emp)
    for d in range(1, days + 1):
        ws.cell(row=row, column=d + 2, value=roster[emp][d])

# CLEAR EXTRA DAYS
for i, emp in enumerate(employees):
    row = i + 4
    for d in range(days + 1, 32):
        ws.cell(row=row, column=d + 2, value="")

output = io.BytesIO()
wb.save(output)
output.seek(0)

st.success("Roster Generated Successfully")
st.download_button(
    "Download Roster",
    output,
    file_name=f"{month}_ROSTER.xlsx"
)
